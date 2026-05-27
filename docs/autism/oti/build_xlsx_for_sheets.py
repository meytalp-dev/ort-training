# -*- coding: utf-8 -*-
"""
בונה קובץ XLSX מ-all-enriched.json — מסודר לפי קטגוריה ותת-קטגוריה.
טאב 1: "הכול" — כל 842 המענים
טאבים 2-9: אחד לכל קטגוריה
פלט: c:/Users/meyta/Downloads/oti-services.xlsx
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = ROOT / "data" / "all-enriched.json"
OUT = Path(r"c:/Users/meyta/Downloads/oti-services.xlsx")

COLUMNS = [
    "שם",
    "קטגוריה",
    "תת-קטגוריה",
    "גילאים",
    "אזור",
    "טלפון",
    "אתר",
    "התמחות",
    "המלצה",
    "ביטחון",
    "תדירות אזכור",
    "הקשר/ציטוט",
    "הערות",
]

HEADER_FILL = PatternFill(start_color="3F7C8C", end_color="3F7C8C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(size=10)

WIDTHS = {
    "שם": 32, "קטגוריה": 18, "תת-קטגוריה": 38, "גילאים": 18,
    "אזור": 22, "טלפון": 18, "אתר": 38, "התמחות": 42,
    "המלצה": 12, "ביטחון": 10, "תדירות אזכור": 10,
    "הקשר/ציטוט": 60, "הערות": 60,
}


def style_sheet(ws, rows_count):
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = WIDTHS.get(name, 20)
    for r in range(2, rows_count + 2):
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="right", vertical="top", wrap_text=True
            )
            ws.cell(row=r, column=c).font = CELL_FONT
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{rows_count + 1}"


def row_from(item):
    return [item.get(col, "") or "" for col in COLUMNS]


def add_sheet(wb, title, items):
    ws = wb.create_sheet(title=title[:31])
    ws.append(COLUMNS)
    for it in items:
        ws.append(row_from(it))
    style_sheet(ws, len(items))


def main():
    data = json.loads(SRC.read_text(encoding="utf-8"))
    print(f"loaded {len(data)} services")

    data_sorted = sorted(
        data,
        key=lambda x: (x.get("קטגוריה", ""), x.get("תת-קטגוריה", ""), x.get("שם", "")),
    )

    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(wb, "הכול", data_sorted)

    by_cat = {}
    for it in data_sorted:
        by_cat.setdefault(it.get("קטגוריה", "ללא קטגוריה"), []).append(it)

    for cat, items in sorted(by_cat.items(), key=lambda kv: -len(kv[1])):
        add_sheet(wb, cat, items)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"saved: {OUT}")
    print(f"tabs: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
