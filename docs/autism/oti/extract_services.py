#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_services.py
-------------------
Scans two WhatsApp chat exports (Hebrew) for autism-related services,
therapists, organizations, programs, frameworks, etc. and produces an
XLSX directory at docs/autism/oti/data/services-extracted.xlsx.

Usage:
    python extract_services.py

Inputs (hard-coded paths):
    /tmp/autism-chats/mothers/chat.txt
    /tmp/autism-chats/jerusalem/chat.txt

Output:
    docs/autism/oti/data/services-extracted.xlsx
        - sheet "הכול" with all entries
        - one sheet per category
"""

import io
import os
import re
import sys
from collections import defaultdict, Counter

# force utf-8 stdout so Hebrew prints on Windows cp1252 consoles
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. run: pip install openpyxl")
    sys.exit(1)


# ---------- paths ----------
def _resolve_chat(rel):
    """Try bash /tmp first, then Windows %TEMP%."""
    candidates = [
        os.path.join("/tmp", rel),
        os.path.join(os.environ.get("TEMP", ""), rel),
        os.path.join(os.environ.get("TMP", ""), rel),
    ]
    for c in candidates:
        if c and os.path.exists(c):
            return c
    return candidates[0]  # fallback, will error nicely


MOTHERS = _resolve_chat("autism-chats/mothers/chat.txt")
JERUSALEM = _resolve_chat("autism-chats/jerusalem/chat.txt")
OUT_XLSX = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "data",
    "services-extracted.xlsx",
)
os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)


# ---------- seed lexicons ----------
# category → list of (pattern, subcategory). pattern is a substring match.
SEEDS = {
    "מסגרות חינוך": [
        ("גן תקשורת", "גן תקשורת"),
        ("גני תקשורת", "גן תקשורת"),
        ("כיתת תקשורת", "כיתת תקשורת"),
        ("כיתות תקשורת", "כיתת תקשורת"),
        ("כיתה תקשורתית", "כיתת תקשורת"),
        ("בית ספר תקשורת", "בית ספר תקשורת"),
        ("ביה\"ס תקשורת", "בית ספר תקשורת"),
        ("תיכון תקשורת", "תיכון תקשורת"),
        ("שילוב", "שילוב"),
        ("ילד משולב", "שילוב"),
        ("סייעת צמודה", "שילוב"),
        ("קשת", "מודל קשת"),
        ("מודל מופת", "מודל מופת"),
        ("פנימייה", "פנימייה"),
        ("מעון רב נכותי", "מעון רב-נכותי"),
        ("מעון רב-נכותי", "מעון רב-נכותי"),
        ("חינוך מיוחד", "חינוך מיוחד"),
        ("מתי\"א", "מתי\"א"),
        ("מתיא", "מתי\"א"),
    ],
    "טיפולים ומטפלים": [
        ("קלינאית תקשורת", "קלינאות תקשורת"),
        ("קלינאי תקשורת", "קלינאות תקשורת"),
        ("ריפוי בעיסוק", "ריפוי בעיסוק"),
        ("מרפאה בעיסוק", "ריפוי בעיסוק"),
        ("מרפא בעיסוק", "ריפוי בעיסוק"),
        ("פסיכולוג", "פסיכולוג"),
        ("פסיכיאטר", "פסיכיאטר"),
        ("מטפל רגשי", "טיפול רגשי"),
        ("מטפלת רגשית", "טיפול רגשי"),
        ("טיפול רגשי", "טיפול רגשי"),
        ("טיפול התנהגותי", "טיפול התנהגותי"),
        ("ABA", "ABA"),
        ("איי בי איי", "ABA"),
        ("PECS", "PECS"),
        ("DIR", "DIR/Floortime"),
        ("Floortime", "DIR/Floortime"),
        ("פלורטיים", "DIR/Floortime"),
        ("אינטגרציה סנסורית", "אינטגרציה סנסורית"),
        ("הידרותרפיה", "הידרותרפיה"),
        ("רכיבה טיפולית", "רכיבה טיפולית"),
        ("טיפול בעזרת בעלי חיים", "טיפול בבע\"ח"),
        ("טיפול בבע\"ח", "טיפול בבע\"ח"),
        ("טיפול באומנות", "טיפול באומנות"),
        ("טיפול בדרמה", "טיפול בדרמה"),
        ("טיפול במוזיקה", "טיפול במוזיקה"),
        ("עובדת סוציאלית", "עו\"ס"),
        ("עו\"ס", "עו\"ס"),
        ("CBT", "CBT"),
        ("מיינדפולנס", "מיינדפולנס"),
    ],
    "אבחון והערכה": [
        ("אבחון דיד", "אבחון"),
        ("אבחון ADOS", "ADOS"),
        ("ADOS", "ADOS"),
        ("מכון להתפתחות הילד", "מכון להתפתחות"),
        ("מכון התפתחות", "מכון להתפתחות"),
        ("התפתחות הילד", "התפתחות הילד"),
        ("ועדת השמה", "ועדת השמה"),
        ("ועדת זכאות", "ועדת זכאות"),
        ("ועדת איפיון", "ועדת איפיון"),
        ("נוירולוג", "נוירולוג"),
        ("אבחון פסיכו-דידקטי", "אבחון פסיכו-דידקטי"),
        ("אבחון פסיכודידקטי", "אבחון פסיכו-דידקטי"),
        ("אבחון רב מקצועי", "אבחון רב מקצועי"),
    ],
    "שירותים לבוגרים": [
        ("דיור נתמך", "דיור נתמך"),
        ("דיור מוגן", "דיור מוגן"),
        ("הוסטל", "הוסטל"),
        ("מרכז יום", "מרכז יום"),
        ("תעסוקה נתמכת", "תעסוקה נתמכת"),
        ("תעסוקה מוגנת", "תעסוקה מוגנת"),
        ("של\"צ", "של\"צ"),
        ("שלצ", "של\"צ"),
        ("שירות לאומי", "שירות לאומי"),
        ("שירות אזרחי", "שירות אזרחי"),
        ("יחידה מיוחדת", "צבא - יחידה מיוחדת"),
        ("רואים רחוק", "צבא - רואים רחוק"),
        ("תוכנית תפקוד", "תוכנית תפקוד"),
        ("אקדמיה", "אקדמיה"),
        ("לוחמים בספקטרום", "צבא - לוחמים בספקטרום"),
        ("בוגרים על הרצף", "בוגרים"),
        ("קהילה תומכת", "קהילה תומכת"),
    ],
    "זכויות ורווחה": [
        ("ביטוח לאומי", "ביטוח לאומי"),
        ("בטל\"א", "ביטוח לאומי"),
        ("בטל\"מ", "ביטוח לאומי"),
        ("קצבת נכות", "קצבת נכות"),
        ("גמלת ילד נכה", "גמלת ילד נכה"),
        ("גימלת נכות", "קצבת נכות"),
        ("סל שיקום", "סל שיקום"),
        ("משרד הרווחה", "משרד הרווחה"),
        ("משרד הבריאות", "משרד הבריאות"),
        ("משרד החינוך", "משרד החינוך"),
        ("כל זכות", "כל זכות"),
        ("kolzchut", "כל זכות"),
        ("אפוטרופסות", "אפוטרופסות"),
        ("ייפוי כוח מתמשך", "ייפוי כוח מתמשך"),
        ("תעודת עיוור", "תעודת נכות"),
        ("תעודת נכה", "תעודת נכות"),
        ("זכאות 91", "זכאות 91"),
        ("סעיף 9", "סעיף 9"),
    ],
    "ארגונים ועמותות": [
        ("אלו\"ט", "אלו\"ט"),
        ("אלוט", "אלו\"ט"),
        ("עמותת אלוט", "אלו\"ט"),
        ("עמותת אלו\"ט", "אלו\"ט"),
        ("עלה נגב", "עלה"),
        ("עמותת עלה", "עלה"),
        ("עמותת אקים", "אקים"),
        ("אקים", "אקים"),
        ("בית איזי שפירא", "בית איזי שפירא"),
        ("איזי שפירא", "בית איזי שפירא"),
        ("בית אקשטיין", "בית אקשטיין"),
        ("אקשטיין", "בית אקשטיין"),
        ("בית נועם", "בית נועם"),
        ("שלוה", "שלוה"),
        ("שלווה", "שלוה"),
        ("אוטיסושיאל", "אוטיסושיאל"),
        ("אוטו סושיאל", "אוטיסושיאל"),
        ("אוטוסושיאל", "אוטיסושיאל"),
        ("איל\"ן", "איל\"ן"),
        ("עזר מציון", "עזר מציון"),
        ("שלו\"ה", "שלוה"),
        ("אופק לילדנו", "אופק לילדנו"),
        ("נאמן", "נאמן"),
        ("תעסוקה שווה", "תעסוקה שווה"),
        ("קרן שלם", "קרן שלם"),
        ("יד שרה", "יד שרה"),
        ("מילבת", "מילבת"),
        ("מילב\"ת", "מילבת"),
        ("אקדם", "אקדם"),
        ("בקשר אמיץ", "בקשר אמיץ"),
        ("אתגרים", "אתגרים"),
        ("צעד קדימה", "צעד קדימה"),
        ("פעמונים", "פעמונים"),
        ("קדימה מדע", "קדימה מדע"),
        ("הרוח הנכונה", "הרוח הנכונה"),
        ("נגישות ישראל", "נגישות ישראל"),
    ],
    "קייטנות ופנאי": [
        ("קייטנה", "קייטנה"),
        ("קייטנת", "קייטנה"),
        ("נופשון", "נופשון"),
        ("נופשונים", "נופשון"),
        ("מחנה קיץ", "מחנה קיץ"),
        ("מחנה חורף", "מחנה חורף"),
        ("חוג חברתי", "חוג חברתי"),
        ("מועדונית", "מועדונית"),
        ("מועדוניות", "מועדונית"),
        ("צופים במיוחד", "תנועות נוער מותאמות"),
        ("כנפיים של קרמבו", "כנפיים של קרמבו"),
        ("קרמבו", "כנפיים של קרמבו"),
        ("אתנחתא", "אתנחתא"),
        ("שיקום בקהילה", "שיקום בקהילה"),
    ],
    "ציוד ועזרים": [
        ("חליפת משקל", "חליפת משקל"),
        ("חליפה סנסורית", "חליפה סנסורית"),
        ("אוזניות מבטלות רעש", "אוזניות מבטלות רעש"),
        ("כדור קפיצה", "ציוד סנסורי"),
        ("טרמפולינה", "ציוד סנסורי"),
        ("גלגל תקשורת", "תקשורת תומכת"),
        ("תמ\"ח", "תקשורת תומכת"),
        ("לוח תקשורת", "תקשורת תומכת"),
        ("Proloquo", "אפליקציית תקשורת"),
        ("Tobii", "מחשב עיניים"),
        ("טובי", "מחשב עיניים"),
        ("אפליקציה לתקשורת", "אפליקציית תקשורת"),
    ],
}

# build inverted index: pattern -> (category, sub)
PATTERN_INDEX = []  # list of (pattern, category, sub), longest first for greedy match
for cat, pairs in SEEDS.items():
    for pat, sub in pairs:
        PATTERN_INDEX.append((pat, cat, sub))
PATTERN_INDEX.sort(key=lambda x: -len(x[0]))


# ---------- canonicalization / alias map ----------
# maps noisy variants → one canonical key (keeps counts coherent)
ALIAS_MAP = {
    "אלוט": "אלו\"ט",
    "עמותת אלוט": "אלו\"ט",
    "עמותת אלו\"ט": "אלו\"ט",
    "עלה נגב": "עלה",
    "עמותת עלה": "עלה",
    "עמותת אקים": "אקים",
    "איזי שפירא": "בית איזי שפירא",
    "אקשטיין": "בית אקשטיין",
    "שלווה": "שלוה",
    "שלו\"ה": "שלוה",
    "אוטו סושיאל": "אוטיסושיאל",
    "אוטוסושיאל": "אוטיסושיאל",
    "מילב\"ת": "מילבת",
    "מתיא": "מתי\"א",
    "בטל\"א": "ביטוח לאומי",
    "בטל\"מ": "ביטוח לאומי",
    "גימלת נכות": "קצבת נכות",
    "קרמבו": "כנפיים של קרמבו",
    "kolzchut": "כל זכות",
    "אבחון פסיכודידקטי": "אבחון פסיכו-דידקטי",
    "מרפא בעיסוק": "ריפוי בעיסוק",
    "מרפאה בעיסוק": "ריפוי בעיסוק",
    "אוטיסושיאל": "אוטיסושיאל",
    "גני תקשורת": "גן תקשורת",
    "כיתות תקשורת": "כיתת תקשורת",
    "כיתה תקשורתית": "כיתת תקשורת",
    "ביה\"ס תקשורת": "בית ספר תקשורת",
    "איי בי איי": "ABA",
    "פלורטיים": "DIR/Floortime",
    "Floortime": "DIR/Floortime",
    "DIR": "DIR/Floortime",
    "שלצ": "של\"צ",
    # plural/singular & form dedup
    "נופשונים": "נופשון",
    "קייטנת": "קייטנה",
    "מועדוניות": "מועדונית",
    "מטפל רגשי": "טיפול רגשי",
    "מטפלת רגשית": "טיפול רגשי",
    "עובדת סוציאלית": "עו\"ס",
    "מרפא בעיסוק": "ריפוי בעיסוק",
    "מרפאה בעיסוק": "ריפוי בעיסוק",
    "קלינאי תקשורת": "קלינאות תקשורת",
    "קלינאית תקשורת": "קלינאות תקשורת",
    "פסיכולוג": "פסיכולוג/ית",
    "פסיכיאטר": "פסיכיאטר/ית",
    "טובי": "Tobii (מחשב עיניים)",
}

# WhatsApp system message prefixes — skip these entirely
SYSTEM_PREFIXES = (
    "\u202aיצרת", "יצרת את הקבוצה", "הצטרף", "הצטרפ", "עזב", "שינית",
    "ההודעות והשיחות", "הסר", "הוסרת", "הסרת", "שינה את תיאור",
    "שינית את תיאור", "חסמת", "ביטלת", "<המדיה הושמטה>",
    "דרת הקבוצה", "דרות הקבוצה", "הקבוצה כך ש",
)

# entities that are too generic/noisy to be useful on their own
BLACKLIST_NAMES = {
    "שילוב",           # too generic, usually a verb not a program
    "ילד משולב",
    "סייעת צמודה",
    "מטפל/ת",
}


# ---------- regex patterns ----------
WHATSAPP_LINE = re.compile(
    r"^(\d{1,2}\.\d{1,2}\.\d{2,4}),\s+(\d{1,2}:\d{2})\s+-\s+(.*)$"
)
PHONE_MOBILE = re.compile(r"(?<!\d)(05\d)[\s-]?(\d{3})[\s-]?(\d{4})(?!\d)")
PHONE_LANDLINE = re.compile(r"(?<!\d)(0[2-4,8-9])[\s-]?(\d{3})[\s-]?(\d{4})(?!\d)")
PHONE_SHORT = re.compile(r"\*\d{4}")
PHONE_INTL = re.compile(r"\+972[\s-]?\d[\s\d-]{7,12}")
URL = re.compile(r"https?://[^\s\u0590-\u05FF]+", re.IGNORECASE)
BARE_DOMAIN = re.compile(
    r"(?<![\w@])((?:[a-z0-9][a-z0-9-]*\.)+(?:co\.il|org\.il|ac\.il|gov\.il|net\.il|com|org))(?![\w@])",
    re.IGNORECASE,
)

# Require the title to end with gershayim/period (so "ד״ר" / "דר'" / "פרופ'"), OR
# "דר " followed by space. This avoids matching "דרת" / "דרות" / "דרכי".
DR_NAME = re.compile(
    r'(ד(?:"|״)ר|דר\'|פרופ(?:\'|")|עו(?:"|״)ס|ד"ר)\s+'
)

# words we never want as the FIRST token after the title (not real names)
NAME_STOPWORDS_FIRST = {
    "הקבוצה", "הכי", "יכולים", "לשלוח", "הודעות", "שלום", "תודה", "בבקשה",
    "שאלה", "עזרה", "אנחנו", "אתם", "בוקר", "ערב", "טוב", "רק", "גם",
    "איך", "מה", "מי", "למה", "איפה", "מתי", "היום", "אתמול", "מחר",
    "היא", "הוא", "הם", "הן", "אני", "אתה", "את", "המפגשים", "מנהל",
    "מנהלת", "מזמינה", "ראש", "המטפל", "המטפלת", "המומחה", "הרופא",
    "הילד", "הילדה", "אמא", "אבא", "דרת", "דרות",
}

# words that we strip from the END of a name (descriptive trailers, not name parts)
NAME_TRAILER_STOPWORDS = {
    "הקבוצה", "המפגשים", "מנהל", "מנהלת", "מזמינה", "ראש", "המטפל",
    "המטפלת", "המומחה", "הרופא", "הילד", "הילדה", "אמא", "אבא",
    "היא", "הוא", "מרכז", "קלינית", "נוספת", "פגישה", "שיחה", "המחלקה",
    "האזורית", "למוגבלויות", "לזוגיות", "זוגיות", "המומחית",
}

CITIES = [
    "ירושלים", "תל אביב", "רעננה", "פתח תקווה", "הרצליה", "רמת גן",
    "חיפה", "באר שבע", "נתניה", "מודיעין", "רחובות", "בית שמש",
    "אשדוד", "אשקלון", "רמלה", "לוד", "יבנה", "גבעת זאב", "קרית ארבע",
    "אפרת", "גוש עציון", "קטמון", "רחביה", "בית הכרם", "תלפיות",
    "גילה", "רמות", "נוה יעקב", "פסגת זאב", "עין כרם", "מבשרת",
    "מעלה אדומים", "חולון", "בת ים", "ראשון לציון", "כפר סבא", "הוד השרון",
    "קרית אונו", "גבעתיים", "רמת השרון", "נהריה", "עכו", "קרית שמונה",
    "צפת", "טבריה", "קרית גת", "דימונה", "ערד", "אילת", "יהוד", "אור יהודה",
]
CITIES_REGEX = re.compile("|".join(map(re.escape, CITIES)))
REGIONS_REGEX = re.compile(r"(צפון|דרום|מרכז|שפלה|גוש דן|ירושלים והסביבה)")

AGE_BUCKETS = [
    # (regex, bucket_label)
    (re.compile(r"\b(גן|גנון|גיל הרך|3 שנים|4 שנים|5 שנים)"), "3-6"),
    (re.compile(r"\b(יסודי|כיתה [אבגדהו]|6-12|7 שנים|8 שנים|9 שנים|10 שנים)"), "6-13"),
    (re.compile(r"\b(חט\"?ב|חטיבת ביניים|כיתה [זחט]|13-15)"), "13-16"),
    (re.compile(r"\b(תיכון|כיתה י\"?[אב]?|כיתה י|15-18|תיכונית)"), "15-18"),
    (re.compile(r"\b(צבא|של\"?צ|שירות לאומי|שירות אזרחי|18\+|בוגר|בוגרים|בוגרת)"), "18+"),
    (re.compile(r"\b(דיור|תעסוקה|הוסטל|מרכז יום|21\+)"), "21+"),
]


# ---------- helpers ----------
def parse_chat(path, source_code):
    """Yield (line_index, text) tuples. Merges continuation lines into previous message."""
    messages = []  # list of dict {idx, date, sender, text}
    current = None
    with open(path, "r", encoding="utf-8") as f:
        for i, raw in enumerate(f, start=1):
            raw = raw.rstrip("\n")
            m = WHATSAPP_LINE.match(raw)
            if m:
                if current:
                    messages.append(current)
                date, time, rest = m.group(1), m.group(2), m.group(3)
                # rest is "sender: text" or system message
                if ":" in rest:
                    sender, _, text = rest.partition(":")
                    sender = sender.strip()
                    text = text.strip()
                else:
                    sender = ""
                    text = rest.strip()
                current = {
                    "idx": i,
                    "date": date,
                    "sender": sender,
                    "text": text,
                    "source": source_code,
                }
            else:
                # continuation line
                if current:
                    current["text"] += " " + raw.strip()
        if current:
            messages.append(current)
    return messages


def normalize_phone(raw):
    digits = re.sub(r"\D", "", raw)
    if digits.startswith("972"):
        digits = "0" + digits[3:]
    if len(digits) == 10 and digits.startswith("05"):
        return f"{digits[:3]}-{digits[3:]}"
    if len(digits) == 9 and digits.startswith("0"):
        return f"{digits[:2]}-{digits[2:]}"
    if len(digits) == 10 and digits.startswith("0"):
        return f"{digits[:3]}-{digits[3:]}"
    return raw.strip()


def extract_phones(text):
    out = set()
    for m in PHONE_MOBILE.finditer(text):
        out.add(normalize_phone("".join(m.groups())))
    for m in PHONE_LANDLINE.finditer(text):
        out.add(normalize_phone("".join(m.groups())))
    for m in PHONE_SHORT.finditer(text):
        out.add(m.group(0))
    for m in PHONE_INTL.finditer(text):
        out.add(normalize_phone(m.group(0)))
    # filter obviously bad ones
    out = {p for p in out if len(re.sub(r"\D", "", p)) >= 4}
    return out


def extract_urls(text):
    out = set()
    for m in URL.finditer(text):
        u = m.group(0).rstrip(").,;:\"'")
        if "chat.whatsapp.com" in u or "bit.ly" in u or "forms.gle" in u or "docs.google.com/forms" in u:
            continue
        out.add(u)
    for m in BARE_DOMAIN.finditer(text):
        d = m.group(1).lower()
        if d.endswith((".com", ".org", ".net")) and "." not in d[:-4]:
            continue  # skip single-label
        if any(bad in d for bad in ("whatsapp.com", "bit.ly", "forms.gle", "gmail.com", "youtube.com", "youtu.be")):
            continue
        out.add(d)
    return out


def extract_locations(text):
    cities = set(CITIES_REGEX.findall(text))
    regions = set(REGIONS_REGEX.findall(text))
    return cities | regions


def infer_age(text):
    buckets = set()
    for rx, lbl in AGE_BUCKETS:
        if rx.search(text):
            buckets.add(lbl)
    return buckets


def canonicalize(name):
    return ALIAS_MAP.get(name, name)


def detect_entities(text):
    """Return list of (canonical_name, category, subcategory) found in text."""
    hits = []
    seen_spans = []
    for pat, cat, sub in PATTERN_INDEX:
        idx = 0
        while True:
            j = text.find(pat, idx)
            if j < 0:
                break
            # skip if overlaps an already-matched span
            overlap = False
            for s, e in seen_spans:
                if not (j + len(pat) <= s or j >= e):
                    overlap = True
                    break
            if not overlap:
                seen_spans.append((j, j + len(pat)))
                hits.append((canonicalize(pat), cat, sub))
            idx = j + len(pat)
    return hits


def extract_named_people(text):
    """Heuristic: 'ד״ר/דר/פרופ/עו״ס' + 1-3 Hebrew words → name entity."""
    people = []
    for m in DR_NAME.finditer(text):
        title = m.group(1)
        tail = text[m.end(): m.end() + 40]
        words = re.findall(r"[\u05D0-\u05EA]{2,}", tail)[:3]
        if not words:
            continue
        if words[0] in NAME_STOPWORDS_FIRST:
            continue
        if not (2 <= len(words[0]) <= 8):
            continue
        # strip trailing descriptive words
        while len(words) > 1 and words[-1] in NAME_TRAILER_STOPWORDS:
            words.pop()
        # must still have at least one name word
        if not words:
            continue
        name = title + " " + " ".join(words)
        name = re.sub(r"\s+", " ", name).strip()
        people.append(name)
    return people


# ---------- main ----------
def main():
    print("reading chats...")
    mothers = parse_chat(MOTHERS, "m")
    jerusalem = parse_chat(JERUSALEM, "j")
    all_msgs = mothers + jerusalem
    print(f"  mothers: {len(mothers)} messages")
    print(f"  jerusalem: {len(jerusalem)} messages")

    # registry[(canonical_name, category)] = dict
    registry = defaultdict(lambda: {
        "name": None,
        "category": None,
        "subcategories": Counter(),
        "ages": Counter(),
        "cities": Counter(),
        "phones": Counter(),
        "urls": Counter(),
        "mentions": 0,
        "contexts": [],   # list of (score, source, idx, snippet)
    })

    # pass 1: seed entities
    for msg in all_msgs:
        text = msg["text"]
        if not text:
            continue
        # skip WhatsApp system messages
        if any(p in text[:40] for p in SYSTEM_PREFIXES):
            continue
        hits = detect_entities(text)
        if not hits:
            continue

        phones = extract_phones(text)
        urls = extract_urls(text)
        cities = extract_locations(text)
        ages = infer_age(text)

        # score the message for "informativeness"
        score = (
            2 * len(phones)
            + 2 * len(urls)
            + len(cities)
            + len(ages)
            + min(len(text) // 80, 3)
        )

        # keep a short snippet (first 220 chars)
        snippet = text.strip()
        if len(snippet) > 220:
            snippet = snippet[:217] + "..."

        source_key = f"{msg['source']}:{msg['idx']}"

        for name, cat, sub in hits:
            key = (name, cat)
            ent = registry[key]
            ent["name"] = name
            ent["category"] = cat
            ent["subcategories"][sub] += 1
            ent["mentions"] += 1
            for p in phones:
                ent["phones"][p] += 1
            for u in urls:
                ent["urls"][u] += 1
            for c in cities:
                ent["cities"][c] += 1
            for a in ages:
                ent["ages"][a] += 1
            ent["contexts"].append((score, source_key, snippet))

    # pass 2: named people — treat as own "טיפולים ומטפלים" unless context says otherwise
    people_registry = defaultdict(lambda: {
        "name": None,
        "category": "טיפולים ומטפלים",
        "subcategories": Counter(),
        "ages": Counter(),
        "cities": Counter(),
        "phones": Counter(),
        "urls": Counter(),
        "mentions": 0,
        "contexts": [],
    })

    for msg in all_msgs:
        text = msg["text"]
        if not text:
            continue
        if any(p in text[:40] for p in SYSTEM_PREFIXES):
            continue
        people = extract_named_people(text)
        if not people:
            continue
        phones = extract_phones(text)
        urls = extract_urls(text)
        cities = extract_locations(text)
        ages = infer_age(text)

        snippet = text.strip()
        if len(snippet) > 220:
            snippet = snippet[:217] + "..."
        source_key = f"{msg['source']}:{msg['idx']}"
        score = 2 * len(phones) + 2 * len(urls) + len(cities) + len(ages)

        # infer professional title/subcategory
        if "פסיכיאטר" in text:
            sub = "פסיכיאטר"
        elif "פסיכולוג" in text:
            sub = "פסיכולוג"
        elif "קלינא" in text:
            sub = "קלינאות תקשורת"
        elif "ריפוי בעיסוק" in text:
            sub = "ריפוי בעיסוק"
        elif "עו\"ס" in text or "עובד" in text:
            sub = "עו\"ס"
        else:
            sub = "מטפל/ת"

        for name in people:
            # normalize: strip trailing whitespace and punctuation
            name = name.strip().rstrip(".,;:")
            if len(name) < 5:
                continue
            ent = people_registry[name]
            ent["name"] = name
            ent["subcategories"][sub] += 1
            ent["mentions"] += 1
            for p in phones:
                ent["phones"][p] += 1
            for u in urls:
                ent["urls"][u] += 1
            for c in cities:
                ent["cities"][c] += 1
            for a in ages:
                ent["ages"][a] += 1
            ent["contexts"].append((score, source_key, snippet))

    # merge people into registry, keyed by (name, category)
    for name, data in people_registry.items():
        # only keep if mentioned >=2 times OR has phone/url info
        if data["mentions"] < 2 and not data["phones"] and not data["urls"]:
            continue
        key = (name, "טיפולים ומטפלים")
        registry[key] = data

    # ---------- build rows ----------
    rows = []
    for (name, cat), ent in registry.items():
        if ent["mentions"] < 1:
            continue
        # blacklist generic noise
        if name in BLACKLIST_NAMES:
            continue
        # require at least 2 mentions OR some metadata (phone/url/city) — loose filter
        if ent["mentions"] < 2 and not (ent["phones"] or ent["urls"] or ent["cities"]):
            continue

        sub = ent["subcategories"].most_common(1)[0][0] if ent["subcategories"] else ""
        ages = " / ".join(sorted(ent["ages"].keys())) if ent["ages"] else ""
        cities = " / ".join([c for c, _ in ent["cities"].most_common(3)])
        phones = " / ".join([p for p, _ in ent["phones"].most_common(2)])
        urls = " / ".join([u for u, _ in ent["urls"].most_common(2)])

        # pick top 1-2 contexts by score
        ctxs = sorted(ent["contexts"], key=lambda x: -x[0])[:2]
        quote = " ||| ".join(c[2] for c in ctxs)
        source = ctxs[0][1] if ctxs else ""

        rows.append({
            "שם": name,
            "קטגוריה": cat,
            "תת-קטגוריה": sub,
            "גילאים": ages,
            "אזור": cities,
            "טלפון": phones,
            "אתר": urls,
            "הקשר/ציטוט": quote,
            "מקור": source,
            "תדירות אזכור": ent["mentions"],
        })

    # sort: by category, then mentions desc
    CAT_ORDER = list(SEEDS.keys())
    rows.sort(key=lambda r: (
        CAT_ORDER.index(r["קטגוריה"]) if r["קטגוריה"] in CAT_ORDER else 99,
        -r["תדירות אזכור"],
    ))

    print(f"built {len(rows)} rows.")

    # ---------- write xlsx ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "הכול"
    ws.sheet_view.rightToLeft = True
    headers = ["שם", "קטגוריה", "תת-קטגוריה", "גילאים", "אזור", "טלפון", "אתר", "הקשר/ציטוט", "מקור", "תדירות אזכור"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0B7A75")
    thin = Side(border_style="thin", color="D4D4D4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_headers(sheet):
        sheet.sheet_view.rightToLeft = True
        sheet.append(headers)
        for col_idx, h in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        sheet.row_dimensions[1].height = 26

    def set_widths(sheet):
        widths = {
            1: 28, 2: 18, 3: 20, 4: 12, 5: 22, 6: 22, 7: 32, 8: 80, 9: 14, 10: 10,
        }
        for col, w in widths.items():
            sheet.column_dimensions[get_column_letter(col)].width = w

    def write_row(sheet, r):
        sheet.append([
            r["שם"], r["קטגוריה"], r["תת-קטגוריה"], r["גילאים"],
            r["אזור"], r["טלפון"], r["אתר"], r["הקשר/ציטוט"],
            r["מקור"], r["תדירות אזכור"],
        ])
        # wrap quote column
        cell = sheet.cell(row=sheet.max_row, column=8)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # "All" sheet
    write_headers(ws)
    for r in rows:
        write_row(ws, r)
    set_widths(ws)
    ws.freeze_panes = "A2"

    # Category sheets
    by_cat = defaultdict(list)
    for r in rows:
        by_cat[r["קטגוריה"]].append(r)

    for cat in CAT_ORDER:
        cat_rows = by_cat.get(cat, [])
        sheet_name = cat[:31]  # Excel sheet name limit
        s = wb.create_sheet(sheet_name)
        write_headers(s)
        for r in cat_rows:
            write_row(s, r)
        set_widths(s)
        s.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")

    # ---------- summary ----------
    print("\n=== summary ===")
    cat_counts = Counter(r["קטגוריה"] for r in rows)
    print(f"Extracted {len(rows)} total entries across {len(cat_counts)} categories.")
    for c, n in cat_counts.most_common():
        print(f"  {c}: {n}")
    top10 = sorted(rows, key=lambda r: -r["תדירות אזכור"])[:10]
    print("\nTop 10 by mentions:")
    for r in top10:
        print(f"  {r['תדירות אזכור']:>4}  [{r['קטגוריה']}] {r['שם']}")


if __name__ == "__main__":
    main()
