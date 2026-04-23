"""
מאחד את כל batch_*.json, מבצע דה-דופ חכם, מיזוג נתונים, ויצוא ל-all.json + XLSX.
"""
import json, glob, re, os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

TMP = 'C:/Users/meyta/AppData/Local/Temp/autism-chats/extracted'
OUT_JSON = 'docs/autism/oti/data/all.json'
OUT_XLSX = 'docs/autism/oti/data/services-ai.xlsx'

CAT_ORDER = [
    'ארגונים ועמותות',
    'מסגרות חינוך',
    'טיפולים ומטפלים',
    'אבחון והערכה',
    'שירותים לבוגרים',
    'זכויות ורווחה',
    'קייטנות ופנאי',
    'ציוד ועזרים',
]

HEB_TITLES = re.compile(r'^(ד["״\']ר|ד"ר|פרופ["״\']?|עו["״\']ס|עו"ד|רב|רבנית|גב|מר)\s+', re.I)
PUNCT = re.compile(r'[.,;:!?"\'\u05f4\u05f3()]+')
WS = re.compile(r'\s+')

def canonicalize(name):
    if not name:
        return ''
    n = str(name).strip()
    # Strip titles for dedup
    n = HEB_TITLES.sub('', n)
    n = PUNCT.sub('', n)
    n = WS.sub(' ', n).strip()
    # Normalize common variants
    replacements = [
        ('אלות', 'אלוט'), ('אלו"ט', 'אלוט'), ('אלו״ט', 'אלוט'),
        ('עמותת ', ''), ('ארגון ', ''),
        ('מרכז ', ''), ('מכון ', ''),
    ]
    nn = n
    for old, new in replacements:
        nn = nn.replace(old, new)
    return nn.lower().strip()


def merge_values(a, b, sep=' / '):
    """Merge two string values, dedup tokens."""
    if not a: return b or ''
    if not b: return a or ''
    tokens = set()
    result = []
    for v in [a, b]:
        for tok in re.split(r'\s*[,/]\s*', str(v)):
            tok = tok.strip()
            key = tok.lower()
            if tok and key not in tokens:
                tokens.add(key)
                result.append(tok)
    return sep.join(result)


def pick_best_quote(a, b):
    """Pick the more informative quote (longer, or with phone/URL)."""
    if not a: return b
    if not b: return a
    def score(q):
        s = len(q)
        if re.search(r'0\d{1,2}[-\s]?\d{7}', q): s += 50
        if 'http' in q.lower() or 'www.' in q.lower(): s += 30
        if 'ממליצ' in q or 'מומלצ' in q: s += 20
        return s
    return a if score(a) > score(b) else b


def merge_entities():
    all_entities = []
    for f in sorted(glob.glob(f'{TMP}/batch_*.json')):
        try:
            data = json.load(open(f, encoding='utf-8'))
            all_entities.extend(data)
        except Exception as e:
            print(f'SKIP {f}: {e}')
    print(f'Loaded {len(all_entities)} raw entries')

    # Group by canonical_name
    groups = {}
    for e in all_entities:
        cn = canonicalize(e.get('canonical_name') or e.get('name', ''))
        if not cn or len(cn) < 3:
            continue
        # Skip pure generic words
        if cn in {'פסיכולוג', 'פסיכיאטר', 'קלינאית', 'מרפא בעיסוק', 'מטפל', 'עובד סוציאלי'}:
            continue
        if cn not in groups:
            groups[cn] = dict(e)
            groups[cn]['sources'] = [e.get('source_ref', '')]
            groups[cn]['mentions'] = 1
        else:
            g = groups[cn]
            g['mentions'] += 1
            g['sources'].append(e.get('source_ref', ''))
            # Merge fields
            g['phone'] = merge_values(g.get('phone', ''), e.get('phone', ''))
            g['url'] = merge_values(g.get('url', ''), e.get('url', ''))
            g['location'] = merge_values(g.get('location', ''), e.get('location', ''))
            g['subcategory'] = merge_values(g.get('subcategory', ''), e.get('subcategory', ''))
            g['specialty'] = merge_values(g.get('specialty') or '', e.get('specialty') or '')
            g['age_range'] = merge_values(g.get('age_range', ''), e.get('age_range', ''))
            # Pick best quote
            g['quote'] = pick_best_quote(g.get('quote', ''), e.get('quote', ''))
            # Sentiment: aggregate to most common
            # (keep first, later we can tally)
            g['notes'] = merge_values(g.get('notes', ''), e.get('notes', ''), sep=' · ')

    merged = list(groups.values())
    print(f'After dedup: {len(merged)} unique entities')

    # Sort: category order, then mentions desc
    def sort_key(x):
        cat = x.get('category', '')
        cat_idx = CAT_ORDER.index(cat) if cat in CAT_ORDER else 99
        return (cat_idx, -x.get('mentions', 0), x.get('name', ''))
    merged.sort(key=sort_key)

    return merged


def to_table_rows(merged):
    """Convert merged entities to flat table rows matching the existing all.json schema."""
    rows = []
    for m in merged:
        sources = m.get('sources', [])
        src = sources[0] if sources else ''
        rows.append({
            'שם': m.get('name', ''),
            'קטגוריה': m.get('category', ''),
            'תת-קטגוריה': m.get('subcategory', ''),
            'גילאים': m.get('age_range', ''),
            'אזור': m.get('location', ''),
            'טלפון': m.get('phone', ''),
            'אתר': m.get('url', ''),
            'הקשר/ציטוט': m.get('quote', ''),
            'מקור': src,
            'תדירות אזכור': str(m.get('mentions', 1)),
            # Extra fields from AI extraction
            'התמחות': m.get('specialty', '') or '',
            'המלצה': m.get('sentiment', ''),
            'ביטחון': m.get('confidence', ''),
            'הערות': m.get('notes', ''),
        })
    return rows


def write_xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    def make_sheet(ws, data):
        if not data:
            return
        headers = list(data[0].keys())
        ws.append(headers)
        for r in data:
            ws.append([r.get(h, '') for h in headers])
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='3F7C8C', end_color='3F7C8C', fill_type='solid')
            cell.alignment = Alignment(horizontal='right')
        ws.freeze_panes = 'A2'
        ws.sheet_view.rightToLeft = True
        # Column widths
        widths = {'שם': 28, 'קטגוריה': 18, 'תת-קטגוריה': 22, 'גילאים': 14, 'אזור': 22,
                  'טלפון': 18, 'אתר': 28, 'הקשר/ציטוט': 60, 'מקור': 10,
                  'תדירות אזכור': 10, 'התמחות': 22, 'המלצה': 12, 'ביטחון': 10, 'הערות': 30}
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths.get(h, 15)

    # "All" sheet
    ws = wb.active
    ws.title = 'הכול'
    make_sheet(ws, rows)

    # Per-category sheets
    for cat in CAT_ORDER:
        cat_rows = [r for r in rows if r['קטגוריה'] == cat]
        if cat_rows:
            ws = wb.create_sheet(cat)
            make_sheet(ws, cat_rows)

    wb.save(OUT_XLSX)
    print(f'Wrote XLSX: {OUT_XLSX}')


def main():
    merged = merge_entities()
    rows = to_table_rows(merged)
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=1)
    print(f'Wrote JSON: {OUT_JSON} ({len(rows)} rows)')
    write_xlsx(rows)

    # Category breakdown
    from collections import Counter
    cats = Counter(r['קטגוריה'] for r in rows)
    print('\nCategory breakdown:')
    for c, n in cats.most_common():
        print(f'  {c}: {n}')


if __name__ == '__main__':
    main()
