import json, re
from openpyxl import load_workbook

HEB = re.compile(r"[\u0590-\u05FF]")
def rev_if_heb(s):
    if not s: return s
    if HEB.search(s):
        # reverse while keeping token order sane: reverse full string
        return s[::-1]
    return s

wb = load_workbook(r"c:\Users\meyta\Downloads\SchoolPupilsList.xlsx")
ws = wb.active

students = []
for row in ws.iter_rows(min_row=2, values_only=True):
    # Actual columns in the xlsx (RTL extraction reversed names block):
    # 0=class 1=given 2=family 3=num 4=tz 5=gender 6=birth 7=country
    # 8=street 9=house 10=apt 11=city 12=p1_phone 13=p1_name 14=p2_phone 15=p2_name 16=reg 17=track
    cls, given, fam, num, tz, gender, birth, country, street, house, apt, city, p1p, p1n, p2p, p2n, reg, track = row
    if not given or not fam:
        continue
    fam_s = str(fam).strip()
    if "משפחה" in fam_s or fam_s == "שם":
        continue
    try:
        int(str(num).strip())
    except:
        continue

    students.append({
        "num": int(str(num).strip()),
        "fam": rev_if_heb(str(fam).strip()),
        "given": rev_if_heb(str(given).strip()),
        "class": rev_if_heb(str(cls or "").strip()),
        "tz": str(tz or "").strip(),
        "gender": rev_if_heb(str(gender or "").strip()),
        "birth": str(birth or "").strip(),
        "street": rev_if_heb(str(street or "").strip()),
        "house": str(house or "").strip(),
        "apt": str(apt or "").strip(),
        "city": rev_if_heb(str(city or "").strip()),
        "p1_name": rev_if_heb(str(p1n or "").strip()),
        "p1_phone": str(p1p or "").strip(),
        "p2_name": rev_if_heb(str(p2n or "").strip()),
        "p2_phone": str(p2p or "").strip(),
    })

# Renumber sequentially and sort by class then family name
def class_key(c):
    order = {"ט'1":91,"ט'2":92,"ט'3":93,"י'1":101,"י'2":102,"י'3":103,
             'י"א1':111,'י"א2':112,'י"א3':113,'י"ב1':121,'י"ב2':122,'י"ב3':123}
    return order.get(c.replace(" ",""), 999)

students.sort(key=lambda s: (class_key(s["class"]), s["fam"]))
for i, s in enumerate(students, 1):
    s["num"] = i

print(f"Total students: {len(students)}")
print(json.dumps(students[:2], ensure_ascii=False, indent=2))

with open(r"c:\Users\meyta\Downloads\ort-presentation-builder\_drafts\students.json", "w", encoding="utf-8") as f:
    json.dump(students, f, ensure_ascii=False, indent=2)
