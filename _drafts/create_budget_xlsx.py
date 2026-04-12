from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# Common styles
header_font = Font(bold=True, size=12, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1A365D')
cell_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
green_fill = PatternFill('solid', fgColor='C6F6D5')
red_fill = PatternFill('solid', fgColor='FED7D7')
yellow_fill = PatternFill('solid', fgColor='FEFCBF')
gray_fill = PatternFill('solid', fgColor='EDF2F7')
purple_fill = PatternFill('solid', fgColor='E9D8FD')

def style_headers(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_align
        cell.border = thin_border

def style_row(ws, row, num_cols, fill=None):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = cell_align
        cell.border = thin_border
        if fill:
            cell.fill = fill

# ===== Sheet 1: Regular Budget =====
ws1 = wb.active
ws1.title = '\u05ea\u05e7\u05e6\u05d9\u05d1 \u05e8\u05d2\u05d9\u05dc'
ws1.sheet_view.rightToLeft = True

headers1 = ['\u05ea\u05d0\u05e8\u05d9\u05da', '\u05e1\u05d5\u05d2', '\u05e9\u05e2\u05d5\u05ea', '\u05e9\u05d5\u05d5\u05d9 \u05e9"\u05d7', '\u05e9\u05d5\u05dc\u05d7/\u05de\u05e7\u05d5\u05e8', '\u05e4\u05e8\u05d8\u05d9\u05dd', '\u05e1\u05d8\u05d8\u05d5\u05e1']
ws1.append(headers1)
style_headers(ws1, 7)

# Income rows
rows1 = [
    ["\u05e1\u05e4\u05d8' 2025", '\u05d4\u05db\u05e0\u05e1\u05d4', 634, 3804000, '\u05ea\u05e7\u05e6\u05d9\u05d1 \u05e9\u05e0\u05ea\u05d9 \u05d0\u05d5\u05e8\u05d8', '\u05ea\u05e7\u05df (624) + \u05d0\u05e4\u05e7\u05d8\u05d9\u05d1\u05d9\u05d5\u05ea (10)', '\u05de\u05d0\u05d5\u05e9\u05e8'],
    ["\u05e1\u05e4\u05d8' 2025", '\u05d4\u05db\u05e0\u05e1\u05d4', 37, 222000, '\u05ea\u05d5\u05e1\u05e4\u05ea\u05d9\u05d5\u05ea \u05de\u05e9\u05d4"\u05e2 + \u05d7\u05e0"\u05de', '34 \u05e9"\u05e9 + 3 \u05e9"\u05d1', '\u05de\u05d0\u05d5\u05e9\u05e8'],
    ['19.10.25', '\u05d4\u05db\u05e0\u05e1\u05d4', 50, 300000, '\u05e9\u05dc\u05d5\u05de\u05d9\u05ea \u05d0\u05d5\u05e8\u05e0\u05e9\u05d8\u05d9\u05d9\u05df', '\u05e4\u05e2\u05d9\u05de\u05d4 \u05e8\u05d0\u05e9\u05d5\u05e0\u05d4 \u05ea\u05d5\u05e1\u05e4\u05ea\u05d9\u05d5\u05ea', '\u05de\u05d0\u05d5\u05e9\u05e8'],
    ['8.2.26', '\u05d4\u05db\u05e0\u05e1\u05d4', 80, 480000, '\u05e8\u05d5\u05d9\u05d8\u05dc \u05d0\u05de\u05d9\u05e8', '\u05d0\u05d9\u05e9\u05d5\u05e8 \u05d4\u05de\u05e8\u05ea \u05e9\u05e2\u05d5\u05ea \u05ea\u05e7\u05df \u2014 \u05e4\u05e2\u05d9\u05dc\u05d5\u05d9\u05d5\u05ea \u05e0\u05d5\u05e1\u05e4\u05d5\u05ea', '\u05de\u05d0\u05d5\u05e9\u05e8'],
    ['', '', '', '', '', '', ''],
    ['10.2.26', '\u05d2\u05e8\u05d9\u05e2\u05d4', -60, -360000, '\u05d0\u05e4\u05e8\u05ea \u05db\u05d4\u05df / \u05e9\u05dc\u05d5\u05de\u05d9\u05ea', '\u05d9\u05e8\u05d9\u05d3\u05d4 \u05de-205 \u05dc-185 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd', '\u05d1\u05d5\u05e6\u05e2'],
    ['', '', '', '', '', '', ''],
    ['13.1.26', '\u05d4\u05d5\u05e6\u05d0\u05d4', 33, 198000, '\u05de\u05d9\u05d9\u05d8\u05dc \u2192 \u05e2\u05e0\u05ea + \u05d0\u05e4\u05e8\u05ea', '\u05ea\u05db\u05e0\u05d5\u05df \u05ea\u05e7\u05e6\u05d9\u05d1\u05d9 \u05d0\u05e8\u05d2\u05d5\u05e0\u05d9\u05d5\u05ea', '\u05d0\u05d5\u05e9\u05e8'],
    ['25.1.26', '\u05d4\u05d5\u05e6\u05d0\u05d4', 5, 30000, '\u05de\u05d9\u05d9\u05d8\u05dc \u2192 \u05d0\u05e4\u05e8\u05ea', '\u05d4\u05de\u05e8\u05d4 \u05d6\u05de\u05e0\u05d9\u05ea \u2014 \u05e9\u05db\u05e8 \u05dc\u05d9\u05e0\u05d5\u05d9 \u05d1\u05e0\u05d2\'\u05d9', '\u05d0\u05d5\u05e9\u05e8'],
    ['8.2.26', '\u05d4\u05d5\u05e6\u05d0\u05d4', 80, 520000, '\u05de\u05d9\u05d9\u05d8\u05dc \u2192 \u05d0\u05e4\u05e8\u05ea + \u05e9\u05dc\u05d5\u05de\u05d9\u05ea', '\u05d4\u05de\u05e8\u05d4 80 \u05e9"\u05e9 (\u05d4\u05e6\u05d8\u05d9\u05d9\u05d3\u05d5\u05ea, \u05e6\u05d5\u05d5\u05ea \u05e6\u05e2\u05d9\u05e8, \u05e9\u05d9\u05e4\u05d5\u05e6\u05d9\u05dd, \u05d8\u05d9\u05d5\u05dc\u05d9\u05dd, \u05e1\u05de\u05d9\u05e0\u05e8)', '\u05e6\u05d5\u05de\u05e6\u05dd \u05dc-70'],
    ['9.2.26', '\u05d4\u05d5\u05e6\u05d0\u05d4', 70, 420000, '\u05de\u05d9\u05d9\u05d8\u05dc \u2192 \u05e9\u05dc\u05d5\u05de\u05d9\u05ea', '\u05ea\u05d5\u05db\u05e0\u05d9\u05ea \u05de\u05e2\u05d5\u05d3\u05db\u05e0\u05ea: \u05d4\u05e6\u05d8\u05d9\u05d9\u05d3\u05d5\u05ea 100K, \u05d1\u05d9\u05e9\u05d5\u05dc 50K, \u05e6\u05d5\u05d5\u05ea \u05e6\u05e2\u05d9\u05e8 50K, \u05e9\u05d9\u05e4\u05d5\u05e6\u05d9\u05dd 100K, \u05d8\u05d9\u05d5\u05dc\u05d9\u05dd 100K, \u05e1\u05de\u05d9\u05e0\u05e8 20K', '\u05d1\u05d1\u05d3\u05d9\u05e7\u05d4'],
    ['26.2.26', '\u05d4\u05d5\u05e6\u05d0\u05d4', 60, 350000, '\u05de\u05d9\u05d9\u05d8\u05dc \u2192 \u05d0\u05e4\u05e8\u05ea', '60 \u05ea\u05d5\u05e1\u05e4\u05ea\u05d9\u05d5\u05ea: \u05d4\u05e6\u05d8\u05d9\u05d9\u05d3\u05d5\u05ea 100K, \u05e6\u05d5\u05d5\u05ea \u05e6\u05e2\u05d9\u05e8 50K, \u05e9\u05d9\u05e4\u05d5\u05e6\u05d9\u05dd 80K, \u05d8\u05d9\u05d5\u05dc\u05d9\u05dd 100K, \u05e1\u05de\u05d9\u05e0\u05e8 20K', '\u05d7\u05dc\u05e7\u05d9'],
    ['~26.2.26', '\u05d4\u05d5\u05e6\u05d0\u05d4', 30, 180000, '\u05de\u05d9\u05d9\u05d8\u05dc \u2192 \u05d0\u05e4\u05e8\u05ea', '\u05d4\u05e2\u05d1\u05e8\u05ea \u05de\u05e9\u05e8\u05d5\u05ea \u05de\u05ea\u05e7\u05df \u05dc\u05d7\u05e0"\u05de: \u05e7\u05d5\u05e1\u05e7\u05e1 12, \u05d1\u05e0\u05d9\u05de\u05d9\u05df 8, \u05e8\u05d5\u05d1\u05d1\u05e9\u05d9 6, \u05e9\u05d0\u05d6\u05d5 4', '\u05d8\u05d9\u05d5\u05d8\u05d4'],
    ['8.2.26', '\u05d4\u05d5\u05e6\u05d0\u05d4', 5, 30000, '\u05d0\u05e4\u05e8\u05ea \u05db\u05d4\u05df', '\u05d2\u05e8\u05d9\u05e2\u05ea 5 \u05e9"\u05e9 \u05e7\u05d5\u05d3 1750 \u2014 \u05dc\u05d9\u05e0\u05d5\u05d9 \u05d1\u05e0\u05d2\'\u05d9', '\u05d1\u05d5\u05e6\u05e2'],
]

for i, row in enumerate(rows1):
    ws1.append(row)
    r = i + 2
    stype = row[1] if row[1] else ''
    fill = None
    if stype == '\u05d4\u05db\u05e0\u05e1\u05d4': fill = green_fill
    elif stype == '\u05d2\u05e8\u05d9\u05e2\u05d4': fill = red_fill
    elif stype == '\u05d4\u05d5\u05e6\u05d0\u05d4': fill = yellow_fill
    style_row(ws1, r, 7, fill)

# Summary
ws1.append([])
for summary in [
    ['', '\u05e1\u05d4"\u05db \u05d4\u05db\u05e0\u05e1\u05d5\u05ea', '=C2+C3+C4+C5', '=D2+D3+D4+D5', '', '', ''],
    ['', '\u05e1\u05d4"\u05db \u05d2\u05e8\u05d9\u05e2\u05d5\u05ea', '=C7', '=D7', '', '', ''],
    ['', '\u05e1\u05d4"\u05db \u05d4\u05d5\u05e6\u05d0\u05d5\u05ea \u05de\u05d0\u05d5\u05e9\u05e8\u05d5\u05ea', 38, 228000, '', '33 + 5 \u05e9"\u05e9', ''],
    ['', '\u05e1\u05d4"\u05db \u05d4\u05d5\u05e6\u05d0\u05d5\u05ea \u05d1\u05d1\u05d3\u05d9\u05e7\u05d4', '~160', '', '', '70 + 60 + 30 \u05e9"\u05e9', ''],
]:
    ws1.append(summary)
    r = ws1.max_row
    for c in range(1, 8):
        cell = ws1.cell(row=r, column=c)
        cell.font = Font(bold=True, size=11)
        cell.fill = gray_fill
        cell.alignment = cell_align
        cell.border = thin_border

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 25
ws1.column_dimensions['F'].width = 55
ws1.column_dimensions['G'].width = 16

# ===== Sheet 2: Special Education =====
ws2 = wb.create_sheet('\u05e1\u05dc \u05d7\u05d9\u05e0\u05d5\u05da \u05de\u05d9\u05d5\u05d7\u05d3')
ws2.sheet_view.rightToLeft = True

headers2 = ['\u05e1\u05d5\u05d2', '\u05e9\u05e2\u05d5\u05ea', '\u05e2\u05d5\u05d1\u05d3\u05d9\u05dd', '\u05de\u05e7\u05d5\u05e8 \u05ea\u05e7\u05e6\u05d9\u05d1\u05d9', '\u05d4\u05e2\u05e8\u05d5\u05ea']
ws2.append(headers2)
style_headers(ws2, 5)

rows2 = [
    ['\u05e1\u05dc 103 \u05e9\u05d9\u05dc\u05d5\u05d1', 114, '\u2014', '\u05e7\u05d5\u05d3 1750', '\u05e1\u05dc \u05e9\u05d9\u05dc\u05d5\u05d1 \u05de\u05e7\u05d5\u05e8\u05d9'],
    ['\u05ea\u05d5\u05e1\u05e4\u05ea\u05d9\u05d5\u05ea', 50, '\u2014', '\u05e7\u05d5\u05d3 1750', '\u05e4\u05e2\u05d9\u05dc\u05d5\u05ea \u05ea\u05d5\u05e1\u05e4\u05ea\u05d9\u05ea'],
    ['\u05d4\u05e2\u05d1\u05e8\u05d4 \u05de\u05ea\u05e7\u05df', 30, '\u05e7\u05d5\u05e1\u05e7\u05e1, \u05d1\u05e0\u05d9\u05de\u05d9\u05df, \u05e8\u05d5\u05d1\u05d1\u05e9\u05d9, \u05e9\u05d0\u05d6\u05d5', '\u05ea\u05e7\u05df \u2192 \u05d7\u05e0"\u05de', '\u05d8\u05d9\u05d5\u05d8\u05d4 \u05e4\u05d1\u05e8\' 2026'],
    ['\u05e1\u05d4"\u05db \u05e1\u05dc', '=B2+B3+B4', '\u2014', '\u2014', '\u2014'],
    ['', '', '', '', ''],
    ['\u05e0\u05d9\u05e6\u05d5\u05dc: \u05de\u05d9\u05e0\u05d4\u05dc + \u05d0\u05e4\u05e7\u05d8\u05d9\u05d1\u05d9\u05d9\u05dd', 54.5, '\u05d0\u05d3\u05dc\u05e8, \u05d7\u05e8\u05d9\u05e8\u05d9, \u05d0\u05e1\u05e8\u05e3, \u05d3\u05e8\u05e2\u05d9, \u05d8\u05e0\u05e0\u05d1\u05d0\u05d5\u05dd, \u05d2\u05d5\u05dc\u05df', '\u05e1\u05dc \u05e9\u05d9\u05dc\u05d5\u05d1', ''],
    ['\u05e0\u05d9\u05e6\u05d5\u05dc: \u05ea\u05d5\u05e1\u05e4\u05ea\u05d9\u05d5\u05ea', 20, '\u05dc\u05e8\u05d3\u05df, \u05db\u05d4\u05df \u05d0\u05d3\u05d9\u05e8', '\u05de\u05e2\u05d5\u05e8\u05d1', ''],
    ['\u05e0\u05d9\u05e6\u05d5\u05dc: \u05e4\u05e8\u05d5\u05d9\u05e7\u05d8\u05d9\u05dd 51120', 20, '\u05e7\u05dc\u05d9\u05d9\u05df \u05de\u05d0\u05d9, \u05de\u05d2\u05d3\u05dc\u05d5\u05e8, \u05d0\u05dc\u05d5\u05df \u05e8\u05d6', '\u05e1\u05dc \u05e9\u05d9\u05dc\u05d5\u05d1', ''],
    ['\u05e0\u05d9\u05e6\u05d5\u05dc: \u05e9\u05de\u05e2\u05d5\u05df \u05db\u05d4\u05df', 27, '\u05e9\u05de\u05e2\u05d5\u05df \u05db\u05d4\u05df', '\u05d4\u05e1\u05d1\u05d4 \u05de\u05d4\u05d5\u05e8\u05d0\u05d4', '135% \u05de\u05e9\u05e8\u05d4'],
    ['\u05e1\u05d4"\u05db \u05e0\u05d9\u05e6\u05d5\u05dc', '=B7+B8+B9+B10', '\u2014', '\u2014', ''],
    ['\u05d9\u05ea\u05e8\u05d4 \u05e6\u05e4\u05d5\u05d9\u05d4', '=B5-B11', '\u2014', '\u2014', '~435,000 \u05e9"\u05d7'],
]

for i, row in enumerate(rows2):
    ws2.append(row)
    r = i + 2
    label = str(row[0])
    fill = None
    if '\u05e1\u05d4"\u05db' in label or '\u05d9\u05ea\u05e8\u05d4' in label:
        fill = purple_fill
    elif '\u05e0\u05d9\u05e6\u05d5\u05dc' in label:
        fill = yellow_fill
    elif row[0] and row[0] != '':
        fill = green_fill
    style_row(ws2, r, 5, fill)

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 42
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 22

# ===== Sheet 3: Renovations =====
ws3 = wb.create_sheet('\u05e9\u05d9\u05e4\u05d5\u05e6\u05d9\u05dd')
ws3.sheet_view.rightToLeft = True

headers3 = ['\u05ea\u05d0\u05e8\u05d9\u05da', '\u05e4\u05e2\u05d5\u05dc\u05d4', '\u05e1\u05db\u05d5\u05dd', '\u05e1\u05d8\u05d8\u05d5\u05e1']
ws3.append(headers3)
style_headers(ws3, 4)

rows3 = [
    ['31.12.25', '\u05d0\u05d9\u05e9\u05d5\u05e8 \u05ea\u05e7\u05e6\u05d5\u05d1 \u05de\u05e9\u05e8\u05d3 \u05d4\u05e2\u05d1\u05d5\u05d3\u05d4', 91360, '\u05d0\u05d5\u05e9\u05e8'],
    ['11.1.26', '\u05de\u05d9\u05d9\u05d8\u05dc \u05e9\u05dc\u05d7\u05d4 \u05ea\u05d5\u05db\u05e0\u05d9\u05ea \u05e9\u05d9\u05e4\u05d5\u05e6\u05d9\u05dd', '', '\u05d1\u05d5\u05d8\u05dc'],
    ['12.1.26', '\u05d8\u05dc\u05d9\u05d4 \u05d1\u05e8\u05e7\u05d5\u05d1\u05d9\u05e5: \u05dc\u05d4\u05e4\u05e0\u05d5\u05ea \u05dc\u05e6\u05e0\u05e8\u05ea', '', '\u05de\u05de\u05ea\u05d9\u05df'],
    ['11.3.26', '\u05e9\u05dc\u05d5\u05de\u05d9\u05ea \u05e9\u05d2\u05d9\u05d1 \u05de\u05d1\u05e7\u05e9\u05ea \u05ea\u05d5\u05db\u05e0\u05d9\u05ea + \u05d0\u05d9\u05e9\u05d5\u05e8 \u05de\u05e4\u05e7\u05d7', '', '\u05de\u05de\u05ea\u05d9\u05df'],
    ['17.3.26', '\u05de\u05d9\u05d9\u05d8\u05dc: \u05de\u05d7\u05db\u05d4 \u05dc\u05e2\u05dc\u05d5\u05d9\u05d5\u05ea \u05d0\u05d9\u05e0\u05e1\u05d8\u05dc\u05e6\u05d9\u05d4', '', '\u05e4\u05ea\u05d5\u05d7'],
]

status_colors = {
    '\u05d0\u05d5\u05e9\u05e8': green_fill,
    '\u05d1\u05d5\u05d8\u05dc': red_fill,
    '\u05de\u05de\u05ea\u05d9\u05df': yellow_fill,
    '\u05e4\u05ea\u05d5\u05d7': yellow_fill,
}

for i, row in enumerate(rows3):
    ws3.append(row)
    r = i + 2
    for c in range(1, 5):
        cell = ws3.cell(row=r, column=c)
        cell.alignment = cell_align
        cell.border = thin_border
        if c == 4:
            cell.fill = status_colors.get(row[3], PatternFill())

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 48
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 14

# Number formatting for currency columns
for ws in [ws1]:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

for ws in [ws1]:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

wb.save('_drafts/budget-dashboard.xlsx')
print('Excel file created successfully!')
